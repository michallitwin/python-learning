import csv
import openpyxl
import os

for folderName, _, filenames in os.walk(
    "."
):  # Skip non-xlsx files, load the workbook object.
    for plik in filenames:
        if not plik.endswith("xlsx"):
            continue
        pelnasciezka = os.path.join(folderName, plik)
        wb = openpyxl.load_workbook(pelnasciezka)

        for sheetname in wb.sheetnames:
            sheet = wb[sheetname]
            # tutaj zmieniamy nazwe i tak samo koncowke usuwamy starego pliku i nadajemy csv
            nazwa_csv = f"{plik[:-5]}_{sheetname}.csv"
            # roibmy nowa sciezke
            sciezka_csv = os.path.join(folderName, nazwa_csv)
            #
            with open(sciezka_csv, "w", newline="", encoding="utf-8") as plik_csv:
                # tutaj wkladamy pusty plik zeby umial pisac w csv
                outputWriter = csv.writer(plik_csv)
                # tutaj leci po wierszach i kolumnach z excela
                for rowNum in range(1, sheet.max_row + 1):
                    rowData = []  # append each cell to this list
                    # Loop through each cell in the row.
                    for colNum in range(1, sheet.max_column + 1):
                        # Append each cell's data to rowData.
                        wartosc = sheet.cell(row=rowNum, column=colNum).value
                        # dodajemy wartosci z komorek do pustej listy
                        rowData.append(wartosc)
                    outputWriter.writerow(rowData)
